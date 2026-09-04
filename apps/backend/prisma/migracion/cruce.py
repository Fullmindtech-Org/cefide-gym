#!/usr/bin/env python3
"""
cruce.py — Combina Excel + CSV, cruza con DB real (PGDMP binary),
genera seed-faltantes.ts con DNIs secuenciales 0000000001, 0000000002...
"""
import sys, io, re, struct, zlib
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SQL_PATH       = r"C:\Users\Jerem\Desktop\fullmintech\cefide-gym\cefide-postgrescefide-m8bcdc_2026-09-02T19-11-55-934Z.sql"
XLSX_PATH      = r"C:\Users\Jerem\Desktop\fullmintech\cefide-gym\cruce_alumnos_dni_cefide.xlsx"
CSV_PATH       = r"C:\Users\Jerem\Desktop\fullmintech\cefide-gym\docs\alumnos_version8.csv"
SEED_PROD_PATH = r"C:\Users\Jerem\Desktop\fullmintech\cefide-gym\apps\backend\prisma\seed-prod.ts"
OUT_SEED       = r"C:\Users\Jerem\Desktop\fullmintech\cefide-gym\apps\backend\prisma\seed-faltantes.ts"
OUT_PATCH      = r"C:\Users\Jerem\Desktop\fullmintech\cefide-gym\apps\backend\prisma\patch-datos.ts"

ACT_MAP = {
    'APARATOS': 'APARATOS', 'GIMNASIO': 'GIMNASIO', 'AERO': 'AERO',
    'AEROBICS': 'AERO', 'AEROBICA': 'AERO', 'AEROBIC': 'AERO',
    'SPINNING': 'SPINNING', 'YOGA': 'YOGA', 'ESCALADA': 'ESCALADA',
    'BOXEO': 'BOXEO', 'BOX': 'BOXEO', 'CROSS TRAINING': 'CROSS TRAINING',
    'CROSS FIT': 'CROSS TRAINING', 'CROSS': 'CROSS TRAINING',
    'PILATES': 'PILATES', 'FUNCIONAL': 'FUNCIONAL', 'ZUMBA': 'ZUMBA',
    'KINESIOLOGIA': 'KINESIOLOGIA', 'REHABILITACION': 'REHABILITACION',
    'PERSONAL TRAINER': 'PERSONAL TRAINER',
}

def normalize_act(s):
    if not s: return 'APARATOS'
    u = str(s).strip().upper()
    if u in ACT_MAP: return ACT_MAP[u]
    for k, v in ACT_MAP.items():
        if u.startswith(k[:5]): return v
    return 'APARATOS'

def clean_name_part(s):
    """Elimina prefijos binarios y caracteres no imprimibles de un string."""
    if not s: return ''
    # Mantener solo letras, espacios, acentos comunes, guiones y puntos
    cleaned = re.sub(r'[^\w\s\.\-\']', ' ', str(s), flags=re.UNICODE)
    # Colapsar espacios múltiples
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    # Si el resultado tiene muchos caracteres raros (underscore flood), descartar
    ratio_alpha = sum(1 for c in cleaned if c.isalpha()) / max(1, len(cleaned))
    if ratio_alpha < 0.5:
        return ''
    return cleaned.title()

def esc(s):
    return str(s).replace('\\', '\\\\').replace("'", "\\'").replace('\n', ' ').replace('\r', '') if s else ''

def name_key(apellido, nombre):
    """Clave de dedup normalizada."""
    a = re.sub(r'[^a-zA-Z]', '', apellido.upper())
    n = re.sub(r'[^a-zA-Z]', '', nombre.upper())
    return (a[:10], n[:8])  # primeros N chars para tolerar variaciones menores


# ── 1. Extraer DNIs ya en DB desde PGDMP binary ───────────────────────────────
print("[1/4] Extrayendo DNIs del dump PGDMP (binary scan)...")

dnis_en_db = set()
with open(SQL_PATH, 'rb') as f:
    raw = f.read()

# Escanear bloques size-prefix (4 bytes LE) + zlib
zlib_second = {0x01, 0x9C, 0xDA, 0x5E}
blocks_found = 0
for pos in range(0, len(raw) - 8):
    size = struct.unpack_from('<I', raw, pos)[0]
    if 10 <= size <= 200000 and pos + 4 + size <= len(raw):
        if raw[pos+4] == 0x78 and raw[pos+5] in zlib_second:
            try:
                dec = zlib.decompress(raw[pos+4:pos+4+size])
                text = dec.decode('utf-8', errors='replace')
                blocks_found += 1
                for line in text.splitlines():
                    if '\t' not in line:
                        continue
                    for col in line.split('\t'):
                        col = col.strip()
                        if col.isdigit() and 6 <= len(col) <= 9:
                            dnis_en_db.add(col)
            except zlib.error:
                pass

print(f"  Bloques PGDMP decomprimidos: {blocks_found}")
print(f"  DNIs ya en DB: {len(dnis_en_db)}")
if dnis_en_db:
    print(f"  Muestra: {list(dnis_en_db)[:5]}")


# ── 2. Leer CSV alumnos_version8.csv ─────────────────────────────────────────
print("[2/4] Leyendo CSV alumnos_version8.csv...")

# CSV columns (header en línea 3, 0-indexed):
# 0=apellido, 1=nombre, 2=nombre_completo, 3=dni, 4=telefono, 5=ciudad, 6=direccion, 7=actividad
CSV_APELLIDO = 0
CSV_NOMBRE   = 1
CSV_DNI      = 3
CSV_CIUDAD   = 5
CSV_DIREC    = 6
CSV_ACT      = 7

alumnos_csv = {}   # key → {apellido, nombre, actividad, ...}
header_found = False
rows_csv = 0

with open(CSV_PATH, encoding='utf-8', errors='replace') as f:
    for line in f:
        if not header_found:
            if 'apellido' in line and 'nombre' in line:
                header_found = True
            continue
        parts = line.rstrip('\n').split(',')
        if len(parts) < 2:
            continue
        rows_csv += 1

        apellido = clean_name_part(parts[CSV_APELLIDO] if len(parts) > CSV_APELLIDO else '')
        nombre   = clean_name_part(parts[CSV_NOMBRE]   if len(parts) > CSV_NOMBRE   else '')
        actividad = normalize_act(parts[CSV_ACT].strip() if len(parts) > CSV_ACT else '')

        if not apellido or not nombre:
            continue
        if len(apellido) < 2 or len(nombre) < 2:
            continue

        key = name_key(apellido, nombre)
        if key not in alumnos_csv:
            alumnos_csv[key] = {
                'apellido': apellido,
                'nombre':   nombre,
                'actividad': actividad,
                'telefono': '',
                'ciudad':   clean_name_part(parts[CSV_CIUDAD] if len(parts) > CSV_CIUDAD else ''),
                'direccion': clean_name_part(parts[CSV_DIREC]  if len(parts) > CSV_DIREC  else ''),
            }

print(f"  Filas CSV procesadas: {rows_csv} | Únicos válidos: {len(alumnos_csv)}")


# ── 3. Leer Excel cruce_alumnos_dni_cefide.xlsx ───────────────────────────────
print("[3/4] Leyendo Excel 'Listado consolidado'...")

COL_NOMBRE    = 1
COL_DOMICILIO = 2
COL_TEL       = 6
COL_CIUDAD    = 7
COL_FNAC      = 16
COL_DEPORTE   = 19

import openpyxl
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb['Listado consolidado']

# Los del Excel sobreescriben/complementan a los del CSV
alumnos_todos = dict(alumnos_csv)  # empieza con CSV
rows_xlsx = 0
rows_xlsx_skip = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or row[0] is None:
        continue
    rows_xlsx += 1

    nombre_raw = str(row[COL_NOMBRE] or '').strip()
    clean = re.sub(r'\s*[\(/][^)]*[\)/]?\s*$', '', nombre_raw).strip()
    clean = re.sub(r'\s*/\w+$', '', clean).strip()
    parts = clean.split()
    if not parts:
        rows_xlsx_skip += 1
        continue

    apellido = parts[0].title()
    nombre   = ' '.join(parts[1:]).title() if len(parts) > 1 else apellido

    if len(apellido) < 2 or len(nombre) < 2:
        rows_xlsx_skip += 1
        continue

    actividad = normalize_act(str(row[COL_DEPORTE] or '').strip())
    tel       = str(row[COL_TEL]       or '').strip()
    ciudad    = str(row[COL_CIUDAD]    or '').strip().title()
    direc     = str(row[COL_DOMICILIO] or '').strip().title()
    fnac_raw  = row[COL_FNAC]

    # Parsear fechaNacimiento DD/MM/YY
    fnac_iso = None
    if fnac_raw:
        try:
            s = str(fnac_raw).strip()
            d, m, y = s.split('/')
            y = int(y); d = int(d); m = int(m)
            y = (2000 + y) if y <= 25 else (1900 + y)
            fnac_iso = f"{y:04d}-{m:02d}-{d:02d}T00:00:00.000Z"
        except Exception:
            fnac_iso = None

    key = name_key(apellido, nombre)
    # Excel tiene prioridad (sobreescribe CSV)
    alumnos_todos[key] = {
        'apellido':  apellido,
        'nombre':    nombre,
        'actividad': actividad,
        'telefono':  tel,
        'ciudad':    ciudad,
        'direccion': direc,
        'fnac':      fnac_iso,
    }

wb.close()
print(f"  Filas Excel: {rows_xlsx} | Saltadas: {rows_xlsx_skip}")
print(f"  Total únicos combinados (CSV+Excel): {len(alumnos_todos)}")


# ── 4. Excluir alumnos ya en seed-prod.ts ────────────────────────────────────
print("[4/5] Excluyendo alumnos ya en seed-prod.ts...")
import re as _re

seed_prod_keys = set()
with open(SEED_PROD_PATH, encoding='utf-8', errors='replace') as f:
    for line in f:
        m = _re.search(r"nombre:\s*'([^']*)',\s*apellido:\s*'([^']*)'", line)
        if m:
            nombre_sp   = m.group(1).strip()
            apellido_sp = m.group(2).strip()
            seed_prod_keys.add(name_key(apellido_sp, nombre_sp))

antes = len(alumnos_todos)
alumnos_todos = {k: v for k, v in alumnos_todos.items() if k not in seed_prod_keys}
print(f"  En seed-prod: {len(seed_prod_keys)} | Excluidos: {antes - len(alumnos_todos)}")
print(f"  Restantes a importar: {len(alumnos_todos)}")


# ── 5. Asignar DNIs secuenciales y generar seed ───────────────────────────────
print("[5/5] Asignando DNIs secuenciales y generando seed-faltantes.ts...")

# Todos los alumnos combinados son "faltantes" (la DB tiene apenas 3)
faltantes = list(alumnos_todos.values())
print(f"  Total a importar: {len(faltantes)}")

# Asignar DNI secuencial 0000000001, 0000000002, ...
for i, a in enumerate(faltantes, start=1):
    a['dni'] = f"{i:010d}"

actividades_usadas = sorted(set(a['actividad'] for a in faltantes))

alumnos_ts = '\n'.join(
    f"  {{ dni: '{a['dni']}', nombre: '{esc(a['nombre'])}', apellido: '{esc(a['apellido'])}', activo: true }},"
    for a in faltantes
)

insc_ts = '\n'.join(
    f"  {{ dniAlumno: '{a['dni']}', actividad: '{esc(a['actividad'])}' }},"
    for a in faltantes
)

acts_ts = '\n'.join(f"  '{a}'," for a in actividades_usadas)

total = len(faltantes)

seed = f"""/**
 * seed-faltantes.ts — Importa {total} alumnos del sistema viejo.
 * DNIs secuenciales: 0000000001 → {faltantes[-1]['dni'] if faltantes else '0000000000'}
 * Fuentes: alumnos_version8.csv + cruce_alumnos_dni_cefide.xlsx
 *
 * Correr UNA VEZ:
 *   npx ts-node prisma/seed-faltantes.ts
 */
import {{ PrismaClient, Frecuencia }} from '@prisma/client';
const prisma = new PrismaClient();

const ACTIVIDADES = [
{acts_ts}
];

const ALUMNOS: {{ dni: string; nombre: string; apellido: string; activo: boolean }}[] = [
{alumnos_ts}
];

const INSCRIPCIONES: {{ dniAlumno: string; actividad: string }}[] = [
{insc_ts}
];

async function main() {{
  console.log('=== SEED FALTANTES — {total} alumnos ===\\n');

  for (const nombre of ACTIVIDADES) {{
    await prisma.actividad.upsert({{
      where: {{ nombre }}, update: {{}}, create: {{ nombre }},
    }});
  }}

  // Insertar en lotes de 500 para no saturar
  const BATCH = 500;
  let inserted = 0;
  for (let i = 0; i < ALUMNOS.length; i += BATCH) {{
    const batch = ALUMNOS.slice(i, i + BATCH);
    const res = await prisma.alumno.createMany({{ data: batch, skipDuplicates: true }});
    inserted += res.count;
    if (i % 5000 === 0) console.log(`  Alumnos: ${{inserted}}/{total}...`);
  }}
  console.log(`Alumnos insertados: ${{inserted}}`);

  const actMap = new Map(
    await prisma.actividad.findMany({{ select: {{ id: true, nombre: true }} }})
      .then((rows) => rows.map((x) => [x.nombre, x.id]))
  );
  const alumMap = new Map(
    await prisma.alumno.findMany({{ select: {{ id: true, dni: true }} }})
      .then((rows) => rows.map((x) => [x.dni, x.id]))
  );

  let ok = 0, skip = 0;
  for (const insc of INSCRIPCIONES) {{
    const alumnoId    = alumMap.get(insc.dniAlumno);
    const actividadId = actMap.get(insc.actividad);
    if (!alumnoId || !actividadId) {{ skip++; continue; }}
    await prisma.inscripcionActividad.upsert({{
      where: {{ alumnoId_actividadId: {{ alumnoId, actividadId }} }},
      update: {{}},
      create: {{
        alumnoId,
        actividadId,
        frecuencia: Frecuencia.UNA_VEZ,
        clasesTotal: 5,
        clasesUsadas: 0,
        pagado: false,
      }},
    }});
    ok++;
    if (ok % 5000 === 0) console.log(`  Inscripciones: ${{ok}}...`);
  }}
  console.log(`Inscripciones: ${{ok}} creadas, ${{skip}} omitidas`);
  console.log('\\nListo.');
}}

main().catch((e) => {{ console.error(e); process.exit(1); }}).finally(() => prisma.$disconnect());
"""

with open(OUT_SEED, 'w', encoding='utf-8') as f:
    f.write(seed)

print(f"\nArchivo generado: {OUT_SEED}")
print(f"  {total} alumnos | DNIs: 0000000001 → {faltantes[-1]['dni'] if faltantes else 'N/A'}")
print(f"\nCorrer: npx ts-node prisma/seed-faltantes.ts")


# ── 6. Generar patch-datos.ts (telefono + fechaNacimiento desde Excel) ─────────
print("\n[6/6] Generando patch-datos.ts...")

# Leer Excel de nuevo para obtener todos los alumnos con sus datos extra
# (incluyendo los del seed-prod que ya están en DB)
wb2 = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws2 = wb2['Listado consolidado']

patch_entries = []
patch_sin_fnac = 0

for row in ws2.iter_rows(min_row=2, values_only=True):
    if not row or row[0] is None:
        continue
    nombre_raw = str(row[COL_NOMBRE] or '').strip()
    clean = re.sub(r'\s*[\(/][^)]*[\)/]?\s*$', '', nombre_raw).strip()
    clean = re.sub(r'\s*/\w+$', '', clean).strip()
    parts = clean.split()
    if not parts or len(parts[0]) < 2:
        continue

    apellido = parts[0].title()
    nombre   = ' '.join(parts[1:]).title() if len(parts) > 1 else apellido
    tel      = str(row[COL_TEL] or '').strip()
    fnac_raw = row[COL_FNAC]

    fnac_iso = None
    if fnac_raw:
        try:
            s = str(fnac_raw).strip()
            d, m, y = s.split('/')
            y = int(y); d = int(d); m = int(m)
            y = (2000 + y) if y <= 25 else (1900 + y)
            fnac_iso = f"{y:04d}-{m:02d}-{d:02d}T00:00:00.000Z"
        except Exception:
            patch_sin_fnac += 1

    if not fnac_iso and not tel:
        continue

    entry = {'apellido': apellido, 'nombre': nombre}
    if tel:
        entry['telefono'] = tel
    if fnac_iso:
        entry['fnac'] = fnac_iso
    patch_entries.append(entry)

wb2.close()
print(f"  Entradas de parche: {len(patch_entries)} | Sin fechaNac: {patch_sin_fnac}")

def patch_line(e):
    parts = [f"apellido: '{esc(e['apellido'])}'", f"nombre: '{esc(e['nombre'])}'"]
    if e.get('telefono'):
        parts.append(f"telefono: '{esc(e['telefono'])}'")
    if e.get('fnac'):
        parts.append(f"fechaNacimiento: new Date('{e['fnac']}')")
    return "  { " + ", ".join(parts) + " },"

patch_data_ts = '\n'.join(patch_line(e) for e in patch_entries)
patch_total = len(patch_entries)

patch_script = f"""/**
 * patch-datos.ts — Actualiza telefono y fechaNacimiento de alumnos existentes.
 * Fuente: cruce_alumnos_dni_cefide.xlsx
 * {patch_total} alumnos con datos extra.
 *
 * Correr UNA VEZ (después de los seeds):
 *   npx ts-node prisma/patch-datos.ts
 */
import {{ PrismaClient }} from '@prisma/client';
const prisma = new PrismaClient();

type PatchEntry = {{
  apellido: string;
  nombre: string;
  telefono?: string;
  fechaNacimiento?: Date;
}};

const PATCH: PatchEntry[] = [
{patch_data_ts}
];

async function main() {{
  console.log('=== PATCH DATOS — {patch_total} alumnos ===\\n');
  let ok = 0, skip = 0;

  for (const p of PATCH) {{
    const data: Partial<{{ telefono: string; fechaNacimiento: Date }}> = {{}};
    if (p.telefono)        data.telefono        = p.telefono;
    if (p.fechaNacimiento) data.fechaNacimiento = p.fechaNacimiento;
    if (Object.keys(data).length === 0) {{ skip++; continue; }}

    const res = await prisma.alumno.updateMany({{
      where: {{ apellido: p.apellido, nombre: p.nombre }},
      data,
    }});
    if (res.count > 0) ok++;
    else skip++;

    if ((ok + skip) % 2000 === 0) console.log(`  ${{ok + skip}}/{patch_total}...`);
  }}

  console.log(`\\nActualizados: ${{ok}} | Sin match: ${{skip}}`);
  console.log('Listo.');
}}

main().catch((e) => {{ console.error(e); process.exit(1); }}).finally(() => prisma.$disconnect());
"""

with open(OUT_PATCH, 'w', encoding='utf-8') as f:
    f.write(patch_script)

print(f"Archivo generado: {OUT_PATCH}")
print(f"\nCorrer: npx ts-node prisma/patch-datos.ts")
